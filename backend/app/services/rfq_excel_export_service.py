"""RFQ Excel export service.

Creates a new filled copy of an RFQ Excel template, writing values derived from
`/api/v1/rfq/autofill` into the appropriate columns.

Important: never modify the template file in-place.
"""

from __future__ import annotations


# --- Master template layout constants ---
HEADER_ROW = 2        # human-readable column headers
KEY_ROW = 3           # internal field keys (snake_case)
DEFAULT_DATA_ROW = 4  # first row where we write values (append if needed)

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Optional
import logging

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


# Column header comments/tooltips
COLUMN_COMMENTS = {
    "Drilling Time In Min": (
        "Drilling Operation Time\n"
        "━━━━━━━━━━━━━━━━━━━━━━━\n"
        "Time to drill all holes in the part.\n\n"
        "Calculation:\n"
        "• Based on number of holes detected\n"
        "• Considers hole diameter and depth\n"
        "• Typical: 0.5-2 min per hole\n\n"
        "Formula:\n"
        "Time = Σ(Hole Depth / Feed Rate)\n\n"
        "Example:\n"
        "4 holes × 1.5 min = 6 minutes"
    ),
    "Drilling Cost": (
        "Drilling Operation Cost\n"
        "━━━━━━━━━━━━━━━━━━━━━━━\n"
        "Cost for all drilling operations.\n\n"
        "Formula:\n"
        "Cost = Drilling Time × VMC Rate/min\n\n"
        "Example:\n"
        "6 min × ₹7.5/min = ₹45"
    ),
    "Milling Time In Min": (
        "Milling Operation Time\n"
        "━━━━━━━━━━━━━━━━━━━━━━━\n"
        "Time for milling operations:\n"
        "• Slots and keyways\n"
        "• Flat surfaces\n"
        "• Pockets and grooves\n\n"
        "Calculation:\n"
        "• Based on slot/pocket volume\n"
        "• Material removal rate\n\n"
        "Formula:\n"
        "Time = Volume / (Feed × DOC × WOC)\n\n"
        "Example:\n"
        "2 slots × 3 min = 6 minutes"
    ),
    "Milling Cost": (
        "Milling Operation Cost\n"
        "━━━━━━━━━━━━━━━━━━━━━━━\n"
        "Cost for all milling operations.\n\n"
        "Formula:\n"
        "Cost = Milling Time × VMC Rate/min\n\n"
        "Example:\n"
        "6 min × ₹7.5/min = ₹45"
    ),
    "VMC Time In Min": (
        "Total VMC Machine Time\n"
        "━━━━━━━━━━━━━━━━━━━━━━━\n"
        "Combined time on Vertical Machining Center.\n\n"
        "Includes:\n"
        "• Drilling time\n"
        "• Milling time\n"
        "• Setup and tool changes\n\n"
        "Formula:\n"
        "VMC Time = Drilling + Milling"
    ),
    "VMC cost": (
        "Total VMC Cost\n"
        "━━━━━━━━━━━━━━━━━━━━━━━\n"
        "Total cost for VMC operations.\n\n"
        "Formula:\n"
        "VMC Cost = VMC Time × Rate/min\n\n"
        "Typical Rate: ₹7.5/min"
    ),
    "Turning Time In Min": (
        "Turning/Lathe Time\n"
        "━━━━━━━━━━━━━━━━━━━━━━━\n"
        "Time on CNC lathe for:\n"
        "• OD turning (outside diameter)\n"
        "• ID boring (inside diameter)\n"
        "• Facing (end surfaces)\n"
        "• Grooving and threading\n\n"
        "Formula:\n"
        "Time = Finish Length (mm) × 10 / 40\n\n"
        "Example:\n"
        "45mm × 10 / 40 = 11.3 minutes"
    ),
    "Turning cost": (
        "Turning Operation Cost\n"
        "━━━━━━━━━━━━━━━━━━━━━━━\n"
        "Cost for lathe operations.\n\n"
        "Formula:\n"
        "Cost = Turning Time × Rate/min\n\n"
        "Typical Rate: ₹4/min\n\n"
        "Example:\n"
        "11.3 min × ₹4 = ₹45.29"
    ),
    "Roughing Cost": (
        "Roughing Operation Cost\n"
        "━━━━━━━━━━━━━━━━━━━━━━━\n"
        "Initial heavy cuts to remove bulk material.\n\n"
        "Formula:\n"
        "Cost = Finish Length (mm) × 1.5\n\n"
        "Example:\n"
        "45mm × 1.5 = ₹67.50"
    ),
}


# Data validation options (Top 5 currencies, simplified dropdowns)
VALIDATION_OPTIONS = {
    "RFQ Type": ["New", "Repeat", "Requote", "Engineering Change"],
    "RFQ Status": ["Open", "Quoted", "Won", "Lost", "Hold"],
    "Currency": ["USD", "EUR", "GBP", "JPY", "INR"],
    "Part Type": ["Turned", "Milled", "Turned + Milled", "Assembly"],
    "Part Category": ["Small", "Medium", "Large", "Precision"],
}


def _apply_column_comments(ws, header_row: int, dynamic_comments: dict = None) -> None:
    """Apply explanatory comments to column headers.
    
    Args:
        ws: Worksheet
        header_row: Row number containing headers
        dynamic_comments: Optional dict of {header: comment_text} to override/add
    """
    # Build header map
    headers = {}
    for col in range(1, (ws.max_column or 1) + 1):
        h = ws.cell(header_row, col).value
        if h:
            headers[str(h).strip()] = col
    
    # Merge static and dynamic comments
    all_comments = dict(COLUMN_COMMENTS)
    if dynamic_comments:
        all_comments.update(dynamic_comments)
    
    # Apply comments
    for header, comment_text in all_comments.items():
        if header in headers:
            col = headers[header]
            cell = ws.cell(header_row, col)
            comment = Comment(comment_text, "RFQ System")
            comment.width = 280
            comment.height = 200
            cell.comment = comment


def _apply_row_formulas(ws, row: int, header_map: Dict[str, int]) -> None:
    """Inject formulas that exactly match the original RFQ-2025-01369 - R1.xlsx pattern.

    Input columns (written as plain values by set_by_header — NOT touched here):
        Sr.No, Part No, Drawing Number, Part Name, Part Revision,
        Qty/MOQ, Annual Potential Qty, RFQ Type, Material Grade, Material Spec,
        Coating Spec, Special Process, Special Machining Process,
        Finish OD (Inch), Finish ID (Inch), Finish Length (Inch),
        RM OD (Inch), RM ID (Inch), RM Rate,
        VMC Time In Min, Special Process Cost, Others,
        Inspection and testing Cost, Currency, Exchange Rate,
        RFQ Status, Part Type, Part Category.

    Formula columns written here (exact patterns from original xlsx):
        Finish OD MM   = N*25.4
        Finish ID MM   = P*25.4
        Finish Len MM  = R*25.4
        Length (Inch)  = R+0.35               [RM stock length allowance]
        RM Weight Kg   = ((T*25.4)^2*V_mm*0.785*7.86/1M) minus bore
        Material Cost  = X*W
        Roughing Cost  = S*1.5
        Turning Time   = (S*10/40)
        Turning cost   = AA*7.5
        VMC cost       = AB*7.5
        Sub Total      = SUM(AG+AF+AE+AD+AC+Z+Y)
        P&F            = AH*3%
        OH & Profit    = AH*15%
        Rejection Cost = AH*2%
        Price/Each INR = SUM(AK+AJ+AI+AH)
        Price Currency = (AL/AQ)
        Annual Pot.    = AM*G
        RM Contrib %   = Y/AL*100
        MOQ Cost       = AM * Qty/MOQ
    """
    def cl(header: str) -> Optional[str]:
        """Return column letter only (e.g. 'N') for a header, or None if missing."""
        col = header_map.get(_norm(header))
        return get_column_letter(col) if col else None

    def set_formula(header: str, formula: str) -> None:
        col = header_map.get(_norm(header))
        if col is None:
            return
        cell = ws.cell(row, col)
        # Always overwrite — we are the authoritative formula source for this row
        cell.value = formula
        cell.font = STANDARD_FONT
        cell.alignment = STANDARD_ALIGNMENT
        cell.border = BOLD_BORDER

    r = row

    # ── Column letter refs (resolved dynamically — order-independent) ─────────
    N  = cl("Finish OD (Inch)")
    P  = cl("Finish ID (Inch)")
    R  = cl("Finish Length (Inch)")
    T  = cl("RM OD (Inch)")
    U  = cl("RM ID (Inch)")
    V  = cl("Length (Inch)")          # RM Length col — formula written below
    W  = cl("RM Weight Kg")
    X  = cl("RM Rate")
    Y  = cl("Material Cost")
    Z  = cl("Roughing Cost")
    S  = cl("Finish Length (MM)")     # formula written below
    AA = cl("Turning Time In Min")
    AB = cl("VMC Time In Min")
    AC = cl("Turning cost")
    AD = cl("VMC cost")
    AE = cl("Special Process Cost")
    AF = cl("Others")
    AG = cl("Inspection and testing Cost")
    AH = cl("Sub Total")
    AI = cl("P&F")
    AJ = cl("OH & Profit")
    AK = cl("Rejection Cost")
    AL = cl("Price/Each In INR")
    AM = cl("Price/Each In Currency")
    AQ = cl("Exchange Rate")
    G  = cl("Annual Potential Qty")
    AW = cl("MOQ Cost")
    QM = cl("Qty/MOQ")

    # ── 1. MM conversions (=Col{r}*25.4 — no ROUND, exact original) ──────────
    if N:  set_formula("Finish OD (MM)",     f"={N}{r}*25.4")
    if P:  set_formula("Finish ID (MM)",     f"={P}{r}*25.4")
    if R:  set_formula("Finish Length (MM)", f"={R}{r}*25.4")

    # ── 1b. RM OD / ID auto-fill from Finish dims + machining allowance ───────
    #   RM OD  = Finish OD + 0.1"  (nearest stock above finish size)
    #   RM ID  = Finish ID - 0.05" for hollow parts; 0 for solid (ID=0)
    if N:  set_formula("RM OD (Inch)", f"=ROUND({N}{r}+0.1,3)")
    if P:  set_formula("RM ID (Inch)", f"=IF({P}{r}>0,ROUND(MAX(0,{P}{r}-0.05),3),0)")

    # ── 2. RM Length = Finish Length + 0.35 ──────────────────────────────────
    if R:  set_formula("Length (Inch)", f"={R}{r}+0.35")

    # ── 3. RM Weight — exact formula from original xlsx ───────────────────────
    #   OD part : ((T*25.4)*(T*25.4)*(V*25.4)*0.785*7.86) / 1000000
    #   ID part : ((U*25.4)*(U*25.4)*(V*25.4)*0.785*7.86) / 1000000  [subtracted]
    if T and V and U:
        set_formula("RM Weight Kg",
            f"=((({T}{r}*25.4)*({T}{r}*25.4)*({V}{r}*25.4)*0.785*7.86)/1000000)"
            f"-((({U}{r}*25.4)*({U}{r}*25.4)*({V}{r}*25.4)*0.785*7.86)/1000000)")
    elif T and V:
        set_formula("RM Weight Kg",
            f"=((({T}{r}*25.4)*({T}{r}*25.4)*({V}{r}*25.4)*0.785*7.86)/1000000)")

    # ── 4. Material Cost = RM Rate × RM Weight ───────────────────────────────
    if X and W:
        set_formula("Material Cost", f"={X}{r}*{W}{r}")

    # ── 5. Roughing Cost = Finish Length (MM) × 1.5 ──────────────────────────
    if S:  set_formula("Roughing Cost", f"={S}{r}*1.5")

    # ── 6. Turning Time = (Finish Length MM × 10 / 40) ───────────────────────
    if S:  set_formula("Turning Time In Min", f"=({S}{r}*10/40)")

    # ── 7. Turning Cost = Turning Time × 7.5 ─────────────────────────────────
    if AA: set_formula("Turning cost", f"={AA}{r}*7.5")

    # ── 8. VMC Cost = VMC Time × 7.5 ─────────────────────────────────────────
    if AB: set_formula("VMC cost", f"={AB}{r}*7.5")

    # ── 9. Sub Total = SUM(AG+AF+AE+AD+AC+Z+Y)  [original order] ─────────────
    sub_parts = [c for c in [AG, AF, AE, AD, AC, Z, Y] if c]
    if sub_parts:
        set_formula("Sub Total",
            f"=(SUM({'+'.join(f'{c}{r}' for c in sub_parts)}))")

    # ── 10. Markups — exact percentages from original (3 / 15 / 2 %) ─────────
    if AH:
        set_formula("P&F",            f"={AH}{r}*3%")
        set_formula("OH & Profit",    f"={AH}{r}*15%")
        set_formula("Rejection Cost", f"={AH}{r}*2%")

    # ── 11. Price/Each In INR = SUM(AK+AJ+AI+AH) ─────────────────────────────
    inr_parts = [c for c in [AK, AJ, AI, AH] if c]
    if inr_parts:
        set_formula("Price/Each In INR",
            f"=SUM({'+'.join(f'{c}{r}' for c in inr_parts)})")

    # ── 12. Price/Each In Currency = (Price INR / Exchange Rate) ─────────────
    if AL and AQ:
        set_formula("Price/Each In Currency", f"=({AL}{r}/{AQ}{r})")

    # ── 13. Annual Potential = Price Currency × Annual Potential Qty ──────────
    if AM and G:
        set_formula("Annual Potential", f"={AM}{r}*{G}{r}")

    # ── 14. RM Contribution % = Material Cost / Price INR × 100 ──────────────
    if Y and AL:
        set_formula("RM Contribution %", f"={Y}{r}/{AL}{r}*100")

    # ── 15. MOQ Cost = Price/Each In Currency × Qty/MOQ ─────────────────────────
    if AM and QM:
        set_formula("MOQ Cost", f"={AM}{r}*{QM}{r}")
    elif AM:
        set_formula("MOQ Cost", f"={AM}{r}")


def _apply_data_validations(ws, header_row: int, data_start: int = 8, data_end: int = 200) -> None:
    """Apply dropdown validations based on actual column positions."""
    # Build header map from actual positions
    headers = {}
    for col in range(1, (ws.max_column or 1) + 1):
        h = ws.cell(header_row, col).value
        if h:
            headers[str(h).strip()] = col
    
    # Remove existing validations to avoid duplicates
    ws.data_validations.dataValidation = []
    
    # Add validations at correct positions
    for header, options in VALIDATION_OPTIONS.items():
        if header in headers:
            col = headers[header]
            col_letter = get_column_letter(col)
            
            dv = DataValidation(
                type="list",
                formula1=f'"{",".join(options)}"',
                allow_blank=True
            )
            dv.error = f"Select from: {', '.join(options)}"
            dv.errorTitle = f"Invalid {header}"
            dv.prompt = f"Select {header}"
            dv.promptTitle = header
            
            dv.add(f"{col_letter}{data_start}:{col_letter}{data_end}")
            ws.add_data_validation(dv)


def _create_glossary_sheet(wb) -> None:
    """Create a GLOSSARY sheet with definitions of all RFQ terms."""
    
    # Check if glossary sheet already exists
    if "GLOSSARY" in wb.sheetnames:
        del wb["GLOSSARY"]
    
    # Create new glossary sheet
    glossary_ws = wb.create_sheet("GLOSSARY")
    
    # Define glossary entries: (Term, Category, Definition, Formula/Example)
    glossary_data = [
        # Dimensions
        ("Finish OD", "Dimensions", "Final Outside Diameter of the finished part after all machining operations", "Measured in inches"),
        ("Finish ID", "Dimensions", "Final Inside Diameter (bore) of the finished part", "Measured in inches; 0 for solid parts"),
        ("Finish Length", "Dimensions", "Final length of the finished part after facing operations", "Measured in inches"),
        ("RM OD", "Dimensions", "Raw Material Outside Diameter - stock bar size to purchase", "Finish OD + Allowance (typically +0.1\")"),
        ("RM ID", "Dimensions", "Raw Material Inside Diameter - for tube stock", "Usually 0 for solid bar stock"),
        ("RM Length", "Dimensions", "Raw Material Length - cut length from stock", "Finish Length + Allowance (typically +0.25\")"),
        
        # Materials & Weight
        ("RM Weight", "Material", "Raw Material Weight in kilograms", "π/4 × (OD² - ID²) × Length × Density"),
        ("RM Rate", "Material", "Raw Material cost rate per kilogram", "Price per kg of the material grade"),
        ("Material Cost", "Material", "Total cost of raw material", "RM Weight × RM Rate"),
        ("Material Grade", "Material", "Specification of the material (e.g., 4140, 1018, 65-45-12)", "Steel, Aluminum, Ductile Iron, etc."),
        
        # Machining Operations
        ("Turning Time", "Machining", "Time on CNC lathe for OD/ID turning, facing, grooving", "Finish Length (mm) × 10 / 40"),
        ("Turning Cost", "Machining", "Cost of lathe operations", "Turning Time × Turning Rate (₹4/min typical)"),
        ("Roughing Cost", "Machining", "Cost to remove bulk material before finishing", "Finish Length (mm) × 1.5"),
        ("Drilling Time", "Machining", "Time to drill all holes in the part", "Based on hole count, diameter, and depth"),
        ("Drilling Cost", "Machining", "Cost of drilling operations", "Drilling Time × VMC Rate"),
        ("Milling Time", "Machining", "Time for milling slots, flats, keyways, pockets", "Based on feature volume and removal rate"),
        ("Milling Cost", "Machining", "Cost of milling operations", "Milling Time × VMC Rate"),
        ("VMC Time", "Machining", "Total Vertical Machining Center time", "Drilling Time + Milling Time"),
        ("VMC Cost", "Machining", "Total VMC machine cost", "VMC Time × VMC Rate (₹7.5/min typical)"),
        
        # Cost Components
        ("Sub Total", "Costs", "Sum of all manufacturing costs before overhead", "Material + Roughing + Turning + VMC"),
        ("P&F", "Costs", "Packing & Forwarding charges", "Typically 2-5% of sub total"),
        ("OH & Profit", "Costs", "Overhead and Profit margin", "Typically 15-25% of sub total"),
        ("Rejection Cost", "Costs", "Allowance for rejected parts", "Typically 2-5% for scrap allowance"),
        ("Price/Each INR", "Costs", "Final unit price in Indian Rupees", "Sub Total + P&F + OH + Rejection"),
        ("Price/Each Currency", "Costs", "Final unit price in selected currency", "Price/Each INR ÷ Exchange Rate"),
        
        # Quantities
        ("Qty/MOQ", "Quantity", "Quantity per order / Minimum Order Quantity", "Number of pieces per order"),
        ("Annual Potential Qty", "Quantity", "Expected annual volume", "Yearly demand forecast"),
        ("Annual Potential", "Quantity", "Annual revenue potential", "Price/Each × Annual Qty"),
        ("MOQ Cost", "Quantity", "Cost for minimum order quantity", "Price/Each × MOQ"),
        
        # Contributions
        ("RM Contribution %", "Analysis", "Percentage of price that is raw material", "(Material Cost ÷ Price/Each) × 100"),
        
        # Metadata
        ("Part No", "Metadata", "Unique part number identifier", "Customer part number"),
        ("Drawing Number", "Metadata", "Engineering drawing reference number", "Same as Part No typically"),
        ("Part Name", "Metadata", "Descriptive name of the part", "e.g., PISTON, SHAFT, HOUSING"),
        ("Part Revision", "Metadata", "Drawing revision letter", "A, B, C, etc."),
        ("RFQ Type", "Metadata", "Type of quote request", "New, Repeat, Requote, Engineering Change"),
        ("RFQ Status", "Metadata", "Current status of the RFQ", "Open, Quoted, Won, Lost, Hold"),
        ("Part Type", "Metadata", "Manufacturing method", "Turned, Milled, Turned + Milled, Assembly"),
        ("Part Category", "Metadata", "Size classification", "Small, Medium, Large, Precision"),
        
        # Exchange Rate
        ("Currency", "Financial", "Target currency for pricing", "USD, EUR, GBP, JPY, INR"),
        ("Exchange Rate", "Financial", "Conversion rate to INR", "Live rate updated hourly"),
    ]
    
    # Styles
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    category_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
    term_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)
    wrap_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Set column widths
    glossary_ws.column_dimensions["A"].width = 25
    glossary_ws.column_dimensions["B"].width = 15
    glossary_ws.column_dimensions["C"].width = 55
    glossary_ws.column_dimensions["D"].width = 35
    
    # Title row
    glossary_ws.merge_cells("A1:D1")
    title_cell = glossary_ws["A1"]
    title_cell.value = "RFQ GLOSSARY - Term Definitions"
    title_cell.font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    glossary_ws.row_dimensions[1].height = 30
    
    # Subtitle
    glossary_ws.merge_cells("A2:D2")
    subtitle_cell = glossary_ws["A2"]
    subtitle_cell.value = "Reference guide for understanding RFQ columns and calculations"
    subtitle_cell.font = Font(name="Calibri", size=10, italic=True, color="666666")
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Header row
    headers = ["Term", "Category", "Definition", "Formula / Example"]
    for col, header in enumerate(headers, 1):
        cell = glossary_ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    glossary_ws.row_dimensions[4].height = 25
    
    # Data rows
    current_category = None
    row = 5
    for term, category, definition, formula in glossary_data:
        # Category separator row
        if category != current_category:
            if current_category is not None:
                row += 1  # Add blank row between categories
            current_category = category
        
        # Term
        cell = glossary_ws.cell(row=row, column=1)
        cell.value = term
        cell.font = term_font
        cell.alignment = wrap_align
        cell.border = thin_border
        
        # Category
        cell = glossary_ws.cell(row=row, column=2)
        cell.value = category
        cell.font = normal_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = category_fill
        
        # Definition
        cell = glossary_ws.cell(row=row, column=3)
        cell.value = definition
        cell.font = normal_font
        cell.alignment = wrap_align
        cell.border = thin_border
        
        # Formula/Example
        cell = glossary_ws.cell(row=row, column=4)
        cell.value = formula
        cell.font = Font(name="Consolas", size=10, color="006400")
        cell.alignment = wrap_align
        cell.border = thin_border
        
        glossary_ws.row_dimensions[row].height = 25
        row += 1
    
    # Footer note
    row += 2
    glossary_ws.merge_cells(f"A{row}:D{row}")
    footer_cell = glossary_ws.cell(row=row, column=1)
    footer_cell.value = "Note: Hover over column headers in 'RFQ Details' sheet for quick tooltips. All rates and formulas can be customized."
    footer_cell.font = Font(name="Calibri", size=9, italic=True, color="888888")
    footer_cell.alignment = Alignment(horizontal="center")


# Standard formatting constants
STANDARD_FONT = Font(name="Calibri", size=11)
STANDARD_FONT_BOLD = Font(name="Calibri", size=11, bold=True)
STANDARD_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
BOLD_BORDER = Border(
    left=Side(style="medium"),
    right=Side(style="medium"),
    top=Side(style="medium"),
    bottom=Side(style="medium"),
)


@dataclass(frozen=True)
class ExcelWriteResult:
    output_path: Path
    sheet_name: str
    header_row: int
    row_written: int


def _norm(s: Any) -> str:
    return str(s or "").strip().lower()


def _guess_header_row(ws, max_scan_rows: int = 20) -> int:
    """Pick the row with the most non-empty cells (early rows)."""
    best_row = 1
    best_cnt = -1
    maxr = min(max_scan_rows, ws.max_row or 1)
    for r in range(1, maxr + 1):
        cnt = 0
        for c in range(1, (ws.max_column or 1) + 1):
            v = ws.cell(r, c).value
            if v not in (None, ""):
                cnt += 1
        if cnt > best_cnt:
            best_cnt = cnt
            best_row = r
    return best_row


def _build_header_map(ws, header_row: int) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for c in range(1, (ws.max_column or 1) + 1):
        v = ws.cell(header_row, c).value
        if v in (None, ""):
            continue
        key = _norm(v)
        if key and key not in headers:
            headers[key] = c
    return headers


def _find_or_append_part_row(ws, header_row: int, part_no: str, part_col: int, srno_col: Optional[int]) -> int:
    target = _norm(part_no)
    # Scan existing rows
    for r in range(header_row + 1, (ws.max_row or header_row) + 1):
        v = ws.cell(r, part_col).value
        if _norm(v) == target:
            return r

    # Append new row
    new_r = (ws.max_row or header_row) + 1
    if srno_col:
        # sr no: try to increment from previous numeric
        prev = ws.cell(new_r - 1, srno_col).value
        try:
            ws.cell(new_r, srno_col).value = int(prev) + 1  # type: ignore[arg-type]
        except Exception:
            ws.cell(new_r, srno_col).value = new_r - header_row
    ws.cell(new_r, part_col).value = part_no
    return new_r


def _find_or_create_part_row_master(
    ws,
    header_row: int,
    part_no: str,
    part_col: int,
    srno_col: Optional[int],
    timestamp_col: Optional[int] = None,
) -> tuple:
    """Find existing Part No row or create new one. Returns (row, is_update)."""
    from datetime import datetime
    
    target = _norm(part_no)
    # Scan existing rows
    for r in range(header_row + 1, (ws.max_row or header_row) + 1):
        v = ws.cell(r, part_col).value
        if _norm(v) == target:
            return r, True  # Found existing row, this is an update

    # Append new row
    new_r = (ws.max_row or header_row) + 1
    if srno_col:
        # sr no: try to increment from previous numeric
        prev = ws.cell(new_r - 1, srno_col).value
        try:
            ws.cell(new_r, srno_col).value = int(prev) + 1
        except Exception:
            ws.cell(new_r, srno_col).value = new_r - header_row
    ws.cell(new_r, part_col).value = part_no
    return new_r, False  # New row


def _prepare_comparison_rows(
    ws,
    header_row: int,
    part_no: str,
    part_col: int,
    srno_col: Optional[int],
) -> int:
    """Ensure a comparison row exists below the base row and return the write row."""
    base_row = _find_or_append_part_row(ws, header_row, part_no, part_col, srno_col)
    write_row = base_row + 1
    ws.insert_rows(write_row)
    if srno_col:
        ws.cell(write_row, srno_col).value = None
    ws.cell(write_row, part_col).value = part_no
    return write_row


# Row highlight colors for updates
UPDATE_HIGHLIGHT = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")  # Light yellow
NEW_ROW_HIGHLIGHT = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green


def write_autofill_to_rfq_template(
    *,
    template_path: Path,
    output_path: Path,
    part_no: str,
    autofill_response: Dict[str, Any],
    cost_inputs: Optional[Dict[str, Any]] = None,
    sheet_name: str = "RFQ Details",
) -> ExcelWriteResult:
    """Write autofill values into a copy of the RFQ template.

    `autofill_response` should be the JSON-serializable dict returned by `/rfq/autofill`.
    """
    if not template_path.exists() or not template_path.is_file():
        raise FileNotFoundError(f"Template not found: {template_path}")

    # Load template (never modify on disk in-place)
    wb = openpyxl.load_workbook(template_path, data_only=False)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        used_sheet = sheet_name
    else:
        ws = wb[wb.sheetnames[0]]
        used_sheet = wb.sheetnames[0]

    header_row = HEADER_ROW
    header_map = _build_header_map(ws, header_row=header_row)

    # Required columns by reference sheet
    part_col = header_map.get(_norm("Part No"))
    if not part_col:
        raise ValueError("Could not find 'Part No' column in template")

    srno_col = header_map.get(_norm("Sr.No"))
    row = _prepare_comparison_rows(ws, header_row=header_row, part_no=part_no, part_col=part_col, srno_col=srno_col)

    fields = (autofill_response or {}).get("fields") or {}
    estimate = (autofill_response or {}).get("estimate") or None

    def _shift_header_map(start_col: int, delta: int) -> None:
        for k, v in list(header_map.items()):
            if v >= start_col:
                header_map[k] = v + delta

    def ensure_header_after(header: str, after_header: str) -> int:
        key = _norm(header)
        col = header_map.get(key)
        if col:
            return col
        after_col = header_map.get(_norm(after_header))
        if not after_col:
            raise ValueError(f"Could not find '{after_header}' column in template")
        insert_at = after_col + 1
        ws.insert_cols(insert_at)
        _shift_header_map(insert_at, 1)
        header_cell = ws.cell(header_row, insert_at)
        header_cell.value = header
        # Apply standard formatting to new header cell
        header_cell.font = STANDARD_FONT_BOLD
        header_cell.alignment = STANDARD_ALIGNMENT
        header_cell.border = BOLD_BORDER
        header_map[key] = insert_at
        return insert_at

    def _round_value(value: Any, decimals: int = 2) -> Any:
        """Round numeric values to specified decimal places."""
        if value is None:
            return None
        try:
            num = float(value)
            return round(num, decimals)
        except (TypeError, ValueError):
            return value

    def _apply_cell_formatting(cell) -> None:
        """Apply standard formatting to a cell."""
        cell.font = STANDARD_FONT
        cell.alignment = STANDARD_ALIGNMENT
        cell.border = BOLD_BORDER

    def set_by_header(header: str, value: Any, decimals: int = 2) -> None:
        col = header_map.get(_norm(header))
        if not col:
            return
        rounded_value = _round_value(value, decimals)
        cell = ws.cell(row, col)
        # Preserve any existing formula (user-supplied or copied from template)
        if isinstance(cell.value, str) and cell.value.startswith("="):
            return
        cell.value = rounded_value
        _apply_cell_formatting(cell)

    def fv(field_key: str, decimals: int = 2) -> Optional[float]:
        try:
            v = fields.get(field_key, {}).get("value")
            if v is None:
                return None
            return round(float(v), decimals)
        except Exception:
            return None

    # Log dimension sources before writing to Excel
    finish_od_value = fv("finish_od_in", 3)
    finish_od_source = fields.get("finish_od_in", {}).get("source", "unknown")
    finish_len_value = fv("finish_len_in", 3)
    finish_len_source = fields.get("finish_len_in", {}).get("source", "unknown")
    
    logger.info(f"[Excel Export] Writing dimensions to Excel for part {part_no}:")
    logger.info(f"  - finish_od_in: {finish_od_value} (source: {finish_od_source})")
    logger.info(f"  - finish_len_in: {finish_len_value} (source: {finish_len_source})")
    logger.info(f"  - finish_id_in: {fv('finish_id_in', 3)} (source: {fields.get('finish_id_in', {}).get('source', 'unknown')})")
    logger.info(f"  - rm_od_in: {fv('rm_od_in', 3)} (source: {fields.get('rm_od_in', {}).get('source', 'unknown')})")
    logger.info(f"  - rm_len_in: {fv('rm_len_in', 3)} (source: {fields.get('rm_len_in', {}).get('source', 'unknown')})")
    
    # Check if using envelope values (BUG DETECTION)
    if "envelope" in finish_od_source.lower() or "finish_max_od" in finish_od_source.lower():
        logger.error(f"[Excel Export] ⚠️  BUG DETECTED: finish_od_in is using envelope source: {finish_od_source}")
    if "envelope" in finish_len_source.lower():
        logger.error(f"[Excel Export] ⚠️  BUG DETECTED: finish_len_in is using envelope source: {finish_len_source}")
    
    # Get debug info if available
    debug = (autofill_response or {}).get("debug") or {}
    if debug:
        max_od_debug = debug.get("max_od_in")
        overall_len_debug = debug.get("overall_len_in")
        logger.info(f"[Excel Export] Debug info:")
        logger.info(f"  - debug.max_od_in: {max_od_debug}")
        logger.info(f"  - debug.overall_len_in: {overall_len_debug}")
        logger.info(f"  - debug.od_pool_count: {debug.get('od_pool_count')}")
        logger.info(f"  - debug.used_z_range: {debug.get('used_z_range')}")

    # Dimensions (Inches)
    set_by_header("Finish OD (Inch)", finish_od_value)
    set_by_header("Finish ID (Inch)", fv("finish_id_in", 3))
    set_by_header("Finish Length (Inch)", fv("finish_len_in", 3))
    set_by_header("RM OD (Inch)", fv("rm_od_in", 3))
    set_by_header("RM ID (Inch)", fv("rm_id_in", 3))
    set_by_header("Length (Inch)", fv("rm_len_in", 3))
    
    # Dimensions (MM) — these get overwritten by =col*25.4 formulas anyway
    set_by_header("Finish OD (MM)", fv("finish_od_mm", 3))
    set_by_header("Finish ID (MM)", fv("finish_id_mm", 3))
    set_by_header("Finish Length (MM)", fv("finish_len_mm", 3))

    # Optional: cost inputs (if provided)
    if isinstance(cost_inputs, dict):
        if cost_inputs.get("rm_rate_per_kg") is not None:
            set_by_header("RM Rate", float(cost_inputs["rm_rate_per_kg"]))
        if cost_inputs.get("currency") is not None:
            set_by_header("Currency", str(cost_inputs["currency"]))
        if cost_inputs.get("qty_moq") is not None:
            set_by_header("Qty/MOQ", int(cost_inputs["qty_moq"]))
        if cost_inputs.get("annual_potential_qty") is not None:
            set_by_header("Annual Potential Qty", int(cost_inputs["annual_potential_qty"]))
        
        # Part metadata fields
        if cost_inputs.get("drawing_number"):
            set_by_header("Drawing Number", str(cost_inputs["drawing_number"]))
        if cost_inputs.get("part_name"):
            set_by_header("Part Name", str(cost_inputs["part_name"]))
        if cost_inputs.get("part_revision"):
            set_by_header("Part Revision", str(cost_inputs["part_revision"]))
        if cost_inputs.get("rfq_type"):
            set_by_header("RFQ Type", str(cost_inputs["rfq_type"]))
        if cost_inputs.get("material_grade"):
            set_by_header("Material Grade", str(cost_inputs["material_grade"]))
        if cost_inputs.get("material_spec"):
            set_by_header("Material Spec", str(cost_inputs["material_spec"]))
        if cost_inputs.get("coating_spec"):
            set_by_header("Coating Spec", str(cost_inputs["coating_spec"]))
        if cost_inputs.get("special_process"):
            set_by_header("Special Process", str(cost_inputs["special_process"]))
        if cost_inputs.get("special_machining_process"):
            set_by_header("Special Machining Process", str(cost_inputs["special_machining_process"]))
        
        # Additional metadata fields
        if cost_inputs.get("rfq_status"):
            set_by_header("RFQ Status", str(cost_inputs["rfq_status"]))
        if cost_inputs.get("part_type"):
            set_by_header("Part Type", str(cost_inputs["part_type"]))
        if cost_inputs.get("part_category"):
            set_by_header("Part Category", str(cost_inputs["part_category"]))

    # Optional: estimate block (if present)
    if isinstance(estimate, dict):
        def ev(key: str, decimals: int = 2) -> Optional[float]:
            try:
                v = estimate.get(key, {}).get("value")
                if v is None:
                    return None
                return round(float(v), decimals)
            except Exception:
                return None

        # Weight and material
        set_by_header("RM Weight Kg", ev("rm_weight_kg"))
        set_by_header("Material Cost", ev("material_cost"))
        set_by_header("Roughing Cost", ev("roughing_cost"))
        
        # Turning
        set_by_header("Turning Time In Min", ev("turning_minutes"))
        set_by_header("Turning cost", ev("turning_cost"))
        
        # VMC (includes drilling + milling time)
        set_by_header("VMC Time In Min", ev("vmc_minutes"))
        set_by_header("VMC cost", ev("vmc_cost"))
        # Ensure drilling/milling columns sit between Turning cost and VMC cost
        drill_time_col = ensure_header_after("Drilling Time In Min", "Turning cost")
        ws.cell(row=KEY_ROW, column=drill_time_col).value = "drilling_time_min"
        drill_cost_col = ensure_header_after("Drilling Cost", "Drilling Time In Min")
        ws.cell(row=KEY_ROW, column=drill_cost_col).value = "drilling_cost"
        mill_time_col = ensure_header_after("Milling Time In Min", "Drilling Cost")
        ws.cell(row=KEY_ROW, column=mill_time_col).value = "milling_time_min"
        mill_cost_col = ensure_header_after("Milling Cost", "Milling Time In Min")
        ws.cell(row=KEY_ROW, column=mill_cost_col).value = "milling_cost"
        set_by_header("Drilling Time In Min", ev("drilling_minutes"))
        set_by_header("Drilling Cost", ev("drilling_cost"))
        set_by_header("Milling Time In Min", ev("milling_minutes"))
        set_by_header("Milling Cost", ev("milling_cost"))
        
        # Other costs
        set_by_header("Special Process Cost", ev("special_process_cost"))
        set_by_header("Others", ev("others_cost"))
        set_by_header("Inspection and testing Cost", ev("inspection_cost"))
        
        # Subtotal and markups
        set_by_header("Sub Total", ev("subtotal"))
        set_by_header("P&F", ev("pf_cost"))
        set_by_header("OH & Profit", ev("oh_profit"))
        set_by_header("Rejection Cost", ev("rejection_cost"))
        
        # Final prices
        set_by_header("Price/Each In INR", ev("price_each_inr"))
        set_by_header("Price/Each In Currency", ev("price_each_currency"))
        
        # MOQ Cost formula is handled by _apply_row_formulas (Price/Each × Qty/MOQ)
        # set_by_header here would be overwritten by the formula; skip it.
        
        # Contribution
        set_by_header("RM Contribution %", ev("rm_contribution_pct"))
        
        # Exchange rate (from live API or fallback)
        set_by_header("Exchange Rate", ev("exchange_rate_used"))
        
        # Build dynamic comment for Exchange Rate with source and timestamp
        rate_source = estimate.get("exchange_rate_source", "unknown")
        rate_timestamp = estimate.get("exchange_rate_timestamp", "N/A")
        rate_value = ev("exchange_rate_used")
        currency_used = estimate.get("currency", "USD")
        
        dynamic_exchange_comment = (
            f"Exchange Rate\n"
            f"{'=' * 20}\n\n"
            f"1 {currency_used} = {rate_value} INR\n\n"
            f"Status: {rate_source.upper()}\n"
            f"Last Updated: {rate_timestamp}\n\n"
            f"Rate refreshes every 1 hour."
        )
        
        if rate_source:
            print(f"[Excel Export] Exchange rate source: {rate_source}, timestamp: {rate_timestamp}")
        
        # Annual Potential = Price/Each In Currency × Annual Potential Qty
        set_by_header("Annual Potential", ev("annual_potential"))

    # Apply formatting to header row (bold, centered, bordered)
    for col in range(1, (ws.max_column or 1) + 1):
        header_cell = ws.cell(header_row, col)
        header_cell.font = STANDARD_FONT_BOLD
        header_cell.alignment = STANDARD_ALIGNMENT
        header_cell.border = BOLD_BORDER

    # Highlight the newly written row for easy comparison
    highlight = PatternFill(start_color="E6CCFF", end_color="E6CCFF", fill_type="solid")
    for col in range(1, (ws.max_column or 1) + 1):
        cell = ws.cell(row, col)
        cell.fill = highlight
        # Ensure all cells in the row have standard formatting
        cell.font = STANDARD_FONT_BOLD
        cell.alignment = STANDARD_ALIGNMENT
        cell.border = BOLD_BORDER

    # Inject live Excel formulas so manual edits auto-cascade
    _apply_row_formulas(ws, row, header_map)

    # Apply data validations based on ACTUAL column positions (after any column shifts)
    _apply_data_validations(ws, header_row)
    
    # Apply column comments/tooltips for machining columns
    # dynamic_exchange_comment is built inside the `if estimate:` block above
    dynamic_comments = {"Exchange Rate": dynamic_exchange_comment} if estimate else {}
    _apply_column_comments(ws, header_row, dynamic_comments)
    
    # Add GLOSSARY sheet
    _create_glossary_sheet(wb)

    # Force Excel to fully recalculate all formulas on open
    wb.calculation.calcMode = 'auto'
    wb.calculation.fullCalcOnLoad = True

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return ExcelWriteResult(
        output_path=output_path,
        sheet_name=used_sheet,
        header_row=header_row,
        row_written=row,
    )


def write_autofill_to_master_file(
    *,
    master_path: Path,
    template_path: Path,
    part_no: str,
    autofill_response: Dict[str, Any],
    cost_inputs: Optional[Dict[str, Any]] = None,
    sheet_name: str = "RFQ Details",
) -> ExcelWriteResult:
    """Write/update autofill values in a master RFQ file.
    
    - If master file doesn't exist, create from template
    - If Part No exists, update that row
    - If Part No doesn't exist, append new row
    - Add timestamp column and highlight updated/new rows
    """
    from datetime import datetime
    
    # Create master file from template if it doesn't exist
    if not master_path.exists():
        if not template_path.exists():
            raise FileNotFoundError(f"Template not found: {template_path}")
        master_path.parent.mkdir(parents=True, exist_ok=True)
        # Copy template to master location
        import shutil
        shutil.copy(template_path, master_path)
        print(f"[Excel Master] Created new master file: {master_path.name}")
    
    # Load master file
    wb = openpyxl.load_workbook(master_path, data_only=False)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        used_sheet = sheet_name
    else:
        ws = wb[wb.sheetnames[0]]
        used_sheet = wb.sheetnames[0]

    header_row = HEADER_ROW
    header_map = _build_header_map(ws, header_row=header_row)

    # Required columns
    part_col = header_map.get(_norm("Part No"))
    if not part_col:
        raise ValueError("Could not find 'Part No' column in template")

    srno_col = header_map.get(_norm("Sr.No"))
    
    # Ensure "Last Updated" column exists after Sr.No
    timestamp_col = header_map.get(_norm("Last Updated"))
    if not timestamp_col:
        # Insert timestamp column after Sr.No (or at column 2 if no Sr.No)
        insert_at = (srno_col or 1) + 1
        ws.insert_cols(insert_at)
        # Shift header map
        for k, v in list(header_map.items()):
            if v >= insert_at:
                header_map[k] = v + 1
        # Also update part_col if shifted
        if part_col >= insert_at:
            part_col += 1
        
        header_cell = ws.cell(header_row, insert_at)
        header_cell.value = "Last Updated"
        header_cell.font = STANDARD_FONT_BOLD
        header_cell.alignment = STANDARD_ALIGNMENT
        header_cell.border = BOLD_BORDER
        header_map[_norm("Last Updated")] = insert_at
        timestamp_col = insert_at
        print(f"[Excel Master] Added 'Last Updated' column at position {insert_at}")
    
    
    # Ensure Drilling/Milling headers exist (older templates may miss these columns)
    def ensure_header_after(after_header: str, new_header: str) -> int:
        """Ensure a header exists; if missing, insert a new column immediately after `after_header`.
        Returns the 1-based column index of `new_header`.
        """
        new_key = _norm(new_header)
        if new_key in header_map:
            return header_map[new_key]

        after_key = _norm(after_header)
        after_col = header_map.get(after_key)
        insert_at = (after_col + 1) if after_col else (ws.max_column + 1)

        # Insert column and shift existing header_map indices
        ws.insert_cols(insert_at)
        for k, c in list(header_map.items()):
            if c >= insert_at:
                header_map[k] = c + 1

        # Write header label & copy style from the column to the left (or from itself if appended)
        hdr_cell = ws.cell(row=HEADER_ROW, column=insert_at, value=new_header)
        src_col = insert_at - 1 if insert_at > 1 else insert_at
        src_cell = ws.cell(row=HEADER_ROW, column=src_col)
        try:
            hdr_cell.font = src_cell.font
            hdr_cell.fill = src_cell.fill
            hdr_cell.alignment = src_cell.alignment
            hdr_cell.border = src_cell.border
            hdr_cell.number_format = src_cell.number_format
            hdr_cell.protection = src_cell.protection
        except Exception:
            pass

        header_map[new_key] = insert_at
        print(f"[Excel Master] Inserted missing header '{new_header}' at col {insert_at} (after '{after_header}')")
        return insert_at

    # These columns are required by the master export (even if values are blank)
    after = "VMC cost"
    for hdr in ["Drilling Time In Min", "Drilling Cost", "Milling Time In Min", "Milling Cost"]:
        ensure_header_after(after, hdr)
        after = hdr

# Find or create row for this Part No
    row, is_update = _find_or_create_part_row_master(
        ws, header_row, part_no, part_col, srno_col, timestamp_col
    )
    
    # Set timestamp
    now = datetime.now()
    timestamp_str = now.strftime("%d-%b-%Y %I:%M %p")  # e.g., "31-Jan-2026 09:30 AM"
    ws.cell(row, timestamp_col).value = timestamp_str
    ws.cell(row, timestamp_col).font = Font(name="Calibri", size=9, italic=True)
    ws.cell(row, timestamp_col).alignment = STANDARD_ALIGNMENT
    
    fields = (autofill_response or {}).get("fields") or {}
    estimate = (autofill_response or {}).get("estimate") or None

    def set_by_header(header: str, value: Any, decimals: int = 2) -> None:
        col = header_map.get(_norm(header))
        if not col:
            return
        try:
            if value is not None:
                num = float(value)
                rounded_value = round(num, decimals)
            else:
                rounded_value = None
        except (TypeError, ValueError):
            rounded_value = value
        cell = ws.cell(row, col)
        # Preserve any existing formula (user-supplied or copied from template)
        if isinstance(cell.value, str) and cell.value.startswith("="):
            return
        cell.value = rounded_value
        cell.font = STANDARD_FONT
        cell.alignment = STANDARD_ALIGNMENT
        cell.border = BOLD_BORDER

    def fv(field_key: str, decimals: int = 2) -> Optional[float]:
        try:
            v = fields.get(field_key, {}).get("value")
            if v is None:
                return None
            return round(float(v), decimals)
        except Exception:
            return None

    # Log dimension sources before writing (for master file)
    finish_od_value_master = fv("finish_od_in", 3)
    finish_od_source_master = fields.get("finish_od_in", {}).get("source", "unknown")
    finish_len_value_master = fv("finish_len_in", 3)
    finish_len_source_master = fields.get("finish_len_in", {}).get("source", "unknown")
    
    logger.info(f"[Excel Export - Master] Writing dimensions to Excel for part {part_no}:")
    logger.info(f"  - finish_od_in: {finish_od_value_master} (source: {finish_od_source_master})")
    logger.info(f"  - finish_len_in: {finish_len_value_master} (source: {finish_len_source_master})")
    logger.info(f"  - finish_id_in: {fv('finish_id_in', 3)} (source: {fields.get('finish_id_in', {}).get('source', 'unknown')})")
    logger.info(f"  - rm_od_in: {fv('rm_od_in', 3)} (source: {fields.get('rm_od_in', {}).get('source', 'unknown')})")
    logger.info(f"  - rm_len_in: {fv('rm_len_in', 3)} (source: {fields.get('rm_len_in', {}).get('source', 'unknown')})")
    
    # Check if using envelope values (BUG DETECTION)
    if "envelope" in finish_od_source_master.lower() or "finish_max_od" in finish_od_source_master.lower():
        logger.error(f"[Excel Export - Master] ⚠️  BUG DETECTED: finish_od_in is using envelope source: {finish_od_source_master}")
    if "envelope" in finish_len_source_master.lower():
        logger.error(f"[Excel Export - Master] ⚠️  BUG DETECTED: finish_len_in is using envelope source: {finish_len_source_master}")
    
    # Get debug info if available
    debug_master = (autofill_response or {}).get("debug") or {}
    if debug_master:
        max_od_debug = debug_master.get("max_od_in")
        overall_len_debug = debug_master.get("overall_len_in")
        logger.info(f"[Excel Export - Master] Debug info:")
        logger.info(f"  - debug.max_od_in: {max_od_debug}")
        logger.info(f"  - debug.overall_len_in: {overall_len_debug}")
        logger.info(f"  - debug.od_pool_count: {debug_master.get('od_pool_count')}")
        logger.info(f"  - debug.used_z_range: {debug_master.get('used_z_range')}")

    # Write all the field values (same as before)
    set_by_header("Finish OD (Inch)", finish_od_value_master)
    set_by_header("Finish ID (Inch)", fv("finish_id_in", 3))
    set_by_header("Finish Length (Inch)", finish_len_value_master)
    set_by_header("RM OD (Inch)", fv("rm_od_in", 3))
    set_by_header("RM ID (Inch)", fv("rm_id_in", 3))
    set_by_header("Length (Inch)", fv("rm_len_in", 3))
    set_by_header("Finish OD (MM)", fv("finish_od_mm", 3))
    set_by_header("Finish ID (MM)", fv("finish_id_mm", 3))
    set_by_header("Finish Length (MM)", fv("finish_len_mm", 3))

    # Cost inputs
    if isinstance(cost_inputs, dict):
        if cost_inputs.get("rm_rate_per_kg") is not None:
            set_by_header("RM Rate", float(cost_inputs["rm_rate_per_kg"]))
        if cost_inputs.get("currency") is not None:
            set_by_header("Currency", str(cost_inputs["currency"]))
        if cost_inputs.get("qty_moq") is not None:
            set_by_header("Qty/MOQ", int(cost_inputs["qty_moq"]))
        if cost_inputs.get("annual_potential_qty") is not None:
            set_by_header("Annual Potential Qty", int(cost_inputs["annual_potential_qty"]))
        if cost_inputs.get("drawing_number"):
            set_by_header("Drawing Number", str(cost_inputs["drawing_number"]))
        if cost_inputs.get("part_name"):
            set_by_header("Part Name", str(cost_inputs["part_name"]))
        if cost_inputs.get("part_revision"):
            set_by_header("Part Revision", str(cost_inputs["part_revision"]))
        if cost_inputs.get("rfq_type"):
            set_by_header("RFQ Type", str(cost_inputs["rfq_type"]))
        if cost_inputs.get("material_grade"):
            set_by_header("Material Grade", str(cost_inputs["material_grade"]))
        if cost_inputs.get("rfq_status"):
            set_by_header("RFQ Status", str(cost_inputs["rfq_status"]))
        if cost_inputs.get("part_type"):
            set_by_header("Part Type", str(cost_inputs["part_type"]))

    # Estimate block
    if isinstance(estimate, dict):
        def ev(key: str, decimals: int = 2) -> Optional[float]:
            try:
                v = estimate.get(key, {}).get("value")
                if v is None:
                    return None
                return round(float(v), decimals)
            except Exception:
                return None

        set_by_header("RM Weight Kg", ev("rm_weight_kg"))
        set_by_header("Material Cost", ev("material_cost"))
        set_by_header("Roughing Cost", ev("roughing_cost"))
        set_by_header("Turning Time In Min", ev("turning_minutes"))
        set_by_header("Turning cost", ev("turning_cost"))
        set_by_header("VMC Time In Min", ev("vmc_minutes"))
        set_by_header("VMC cost", ev("vmc_cost"))
        set_by_header("Drilling Time In Min", ev("drilling_minutes"))
        set_by_header("Drilling Cost", ev("drilling_cost"))
        set_by_header("Milling Time In Min", ev("milling_minutes"))
        set_by_header("Milling Cost", ev("milling_cost"))
        set_by_header("Sub Total", ev("subtotal"))
        set_by_header("P&F", ev("pf_cost"))
        set_by_header("OH & Profit", ev("oh_profit"))
        set_by_header("Rejection Cost", ev("rejection_cost"))
        set_by_header("Price/Each In INR", ev("price_each_inr"))
        set_by_header("Price/Each In Currency", ev("price_each_currency"))
        # MOQ Cost formula handled by _apply_row_formulas (Price/Each × Qty/MOQ)
        set_by_header("RM Contribution %", ev("rm_contribution_pct"))
        set_by_header("Exchange Rate", ev("exchange_rate_used"))
        set_by_header("Annual Potential", ev("annual_potential"))

    # Highlight the row based on update/new
    highlight_color = UPDATE_HIGHLIGHT if is_update else NEW_ROW_HIGHLIGHT
    for col in range(1, (ws.max_column or 1) + 1):
        cell = ws.cell(row, col)
        cell.fill = highlight_color
        cell.border = BOLD_BORDER

    # IMPORTANT: Unhide the data row (template may have rows hidden by default)
    ws.row_dimensions[row].hidden = False

    # Inject live Excel formulas so manual edits auto-cascade
    _apply_row_formulas(ws, row, header_map)

    # Apply header formatting
    for col in range(1, (ws.max_column or 1) + 1):
        header_cell = ws.cell(header_row, col)
        header_cell.font = STANDARD_FONT_BOLD
        header_cell.alignment = STANDARD_ALIGNMENT
        header_cell.border = BOLD_BORDER

    # Apply data validations
    _apply_data_validations(ws, header_row)
    
    # Add GLOSSARY sheet if not exists
    if "GLOSSARY" not in wb.sheetnames:
        _create_glossary_sheet(wb)

    # Force Excel to fully recalculate all formulas on open
    wb.calculation.calcMode = 'auto'
    wb.calculation.fullCalcOnLoad = True

    # Save master file
    wb.save(master_path)
    
    action = "Updated" if is_update else "Added"
    print(f"[Excel Master] {action} Part No '{part_no}' at row {row} in {master_path.name}")

    return ExcelWriteResult(
        output_path=master_path,
        sheet_name=used_sheet,
        header_row=header_row,
        row_written=row,
    )


def write_autofill_to_new_file(
    *,
    template_path: Path,
    output_path: Path,
    part_no: str,
    autofill_response: Dict[str, Any],
    cost_inputs: Optional[Dict[str, Any]] = None,
    sheet_name: str = "RFQ Details",
) -> ExcelWriteResult:
    """Create a NEW Excel file from template with only this part's data.
    
    - Creates fresh copy of template
    - Clears any existing data rows (keeps only headers)
    - Adds timestamp column
    - Adds only this part's data as the first row
    """
    from datetime import datetime
    import shutil
    
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")
    
    # Create output directory and copy template
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(template_path, output_path)
    print(f"[Excel NewFile] Created new file: {output_path.name}")
    
    # Load the new file
    wb = openpyxl.load_workbook(output_path, data_only=False)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        used_sheet = sheet_name
    else:
        ws = wb[wb.sheetnames[0]]
        used_sheet = wb.sheetnames[0]

    header_row = HEADER_ROW
    header_map = _build_header_map(ws, header_row=header_row)

    # Required columns
    part_col = header_map.get(_norm("Part No"))
    if not part_col:
        raise ValueError("Could not find 'Part No' column in template")

    srno_col = header_map.get(_norm("Sr.No"))
    
    # Clear ALL rows below header (including placeholder/instruction rows)
    # New data will be written to header_row + 1
    rows_to_delete = (ws.max_row or header_row) - header_row
    if rows_to_delete > 0:
        ws.delete_rows(header_row + 1, rows_to_delete)
        print(f"[Excel NewFile] Deleted {rows_to_delete} rows below header (row {header_row})")
    
    # Data will be written to the row immediately after header
    data_start_row = header_row + 1
    
    # Add "Last Updated" column if not exists
    timestamp_col = header_map.get(_norm("Last Updated"))
    if not timestamp_col:
        insert_at = (srno_col or 1) + 1
        ws.insert_cols(insert_at)
        for k, v in list(header_map.items()):
            if v >= insert_at:
                header_map[k] = v + 1
        if part_col >= insert_at:
            part_col += 1
        
        header_cell = ws.cell(header_row, insert_at)
        header_cell.value = "Last Updated"
        header_cell.font = STANDARD_FONT_BOLD
        header_cell.alignment = STANDARD_ALIGNMENT
        header_cell.border = BOLD_BORDER
        header_map[_norm("Last Updated")] = insert_at
        timestamp_col = insert_at
    
    # Add the new row
    row = data_start_row
    
    # Set Sr.No
    if srno_col:
        ws.cell(row, srno_col).value = 1
    
    # Set Part No
    ws.cell(row, part_col).value = part_no
    
    # Set timestamp
    now = datetime.now()
    timestamp_str = now.strftime("%d-%b-%Y %I:%M %p")
    ws.cell(row, timestamp_col).value = timestamp_str
    ws.cell(row, timestamp_col).font = Font(name="Calibri", size=9, italic=True)
    
    fields = (autofill_response or {}).get("fields") or {}
    estimate = (autofill_response or {}).get("estimate") or None

    def set_by_header(header: str, value: Any, decimals: int = 2) -> None:
        col = header_map.get(_norm(header))
        if not col:
            return
        try:
            if value is not None:
                num = float(value)
                rounded_value = round(num, decimals)
            else:
                rounded_value = None
        except (TypeError, ValueError):
            rounded_value = value
        cell = ws.cell(row, col)
        # Preserve any existing formula (user-supplied or copied from template)
        if isinstance(cell.value, str) and cell.value.startswith("="):
            return
        cell.value = rounded_value
        cell.font = STANDARD_FONT
        cell.alignment = STANDARD_ALIGNMENT
        cell.border = BOLD_BORDER

    def fv(field_key: str, decimals: int = 2) -> Optional[float]:
        try:
            v = fields.get(field_key, {}).get("value")
            if v is None:
                return None
            return round(float(v), decimals)
        except Exception:
            return None

    # Log dimension sources before writing (for new file)
    finish_od_value_new = fv("finish_od_in", 3)
    finish_od_source_new = fields.get("finish_od_in", {}).get("source", "unknown")
    finish_len_value_new = fv("finish_len_in", 3)
    finish_len_source_new = fields.get("finish_len_in", {}).get("source", "unknown")
    
    logger.info(f"[Excel Export - New File] Writing dimensions to Excel for part {part_no}:")
    logger.info(f"  - finish_od_in: {finish_od_value_new} (source: {finish_od_source_new})")
    logger.info(f"  - finish_len_in: {finish_len_value_new} (source: {finish_len_source_new})")
    logger.info(f"  - finish_id_in: {fv('finish_id_in', 3)} (source: {fields.get('finish_id_in', {}).get('source', 'unknown')})")
    logger.info(f"  - rm_od_in: {fv('rm_od_in', 3)} (source: {fields.get('rm_od_in', {}).get('source', 'unknown')})")
    logger.info(f"  - rm_len_in: {fv('rm_len_in', 3)} (source: {fields.get('rm_len_in', {}).get('source', 'unknown')})")
    
    # Check if using envelope values (BUG DETECTION)
    if "envelope" in finish_od_source_new.lower() or "finish_max_od" in finish_od_source_new.lower():
        logger.error(f"[Excel Export - New File] ⚠️  BUG DETECTED: finish_od_in is using envelope source: {finish_od_source_new}")
    if "envelope" in finish_len_source_new.lower():
        logger.error(f"[Excel Export - New File] ⚠️  BUG DETECTED: finish_len_in is using envelope source: {finish_len_source_new}")
    
    # Get debug info if available
    debug_new = (autofill_response or {}).get("debug") or {}
    if debug_new:
        max_od_debug = debug_new.get("max_od_in")
        overall_len_debug = debug_new.get("overall_len_in")
        logger.info(f"[Excel Export - New File] Debug info:")
        logger.info(f"  - debug.max_od_in: {max_od_debug}")
        logger.info(f"  - debug.overall_len_in: {overall_len_debug}")
        logger.info(f"  - debug.od_pool_count: {debug_new.get('od_pool_count')}")
        logger.info(f"  - debug.used_z_range: {debug_new.get('used_z_range')}")

    # Write all field values
    set_by_header("Finish OD (Inch)", finish_od_value_new)
    set_by_header("Finish ID (Inch)", fv("finish_id_in", 3))
    set_by_header("Finish Length (Inch)", finish_len_value_new)
    set_by_header("RM OD (Inch)", fv("rm_od_in", 3))
    set_by_header("RM ID (Inch)", fv("rm_id_in", 3))
    set_by_header("Length (Inch)", fv("rm_len_in", 3))
    set_by_header("Finish OD (MM)", fv("finish_od_mm", 3))
    set_by_header("Finish ID (MM)", fv("finish_id_mm", 3))
    set_by_header("Finish Length (MM)", fv("finish_len_mm", 3))

    # Cost inputs
    if isinstance(cost_inputs, dict):
        if cost_inputs.get("rm_rate_per_kg") is not None:
            set_by_header("RM Rate", float(cost_inputs["rm_rate_per_kg"]))
        if cost_inputs.get("currency") is not None:
            set_by_header("Currency", str(cost_inputs["currency"]))
        if cost_inputs.get("qty_moq") is not None:
            set_by_header("Qty/MOQ", int(cost_inputs["qty_moq"]))
        if cost_inputs.get("annual_potential_qty") is not None:
            set_by_header("Annual Potential Qty", int(cost_inputs["annual_potential_qty"]))
        if cost_inputs.get("drawing_number"):
            set_by_header("Drawing Number", str(cost_inputs["drawing_number"]))
        if cost_inputs.get("part_name"):
            set_by_header("Part Name", str(cost_inputs["part_name"]))
        if cost_inputs.get("part_revision"):
            set_by_header("Part Revision", str(cost_inputs["part_revision"]))
        if cost_inputs.get("rfq_type"):
            set_by_header("RFQ Type", str(cost_inputs["rfq_type"]))
        if cost_inputs.get("material_grade"):
            set_by_header("Material Grade", str(cost_inputs["material_grade"]))
        if cost_inputs.get("rfq_status"):
            set_by_header("RFQ Status", str(cost_inputs["rfq_status"]))
        if cost_inputs.get("part_type"):
            set_by_header("Part Type", str(cost_inputs["part_type"]))

    # Estimate block
    if isinstance(estimate, dict):
        def ev(key: str, decimals: int = 2) -> Optional[float]:
            try:
                v = estimate.get(key, {}).get("value")
                if v is None:
                    return None
                return round(float(v), decimals)
            except Exception:
                return None

        set_by_header("RM Weight Kg", ev("rm_weight_kg"))
        set_by_header("Material Cost", ev("material_cost"))
        set_by_header("Roughing Cost", ev("roughing_cost"))
        set_by_header("Turning Time In Min", ev("turning_minutes"))
        set_by_header("Turning cost", ev("turning_cost"))
        set_by_header("VMC Time In Min", ev("vmc_minutes"))
        set_by_header("VMC cost", ev("vmc_cost"))
        set_by_header("Drilling Time In Min", ev("drilling_minutes"))
        set_by_header("Drilling Cost", ev("drilling_cost"))
        set_by_header("Milling Time In Min", ev("milling_minutes"))
        set_by_header("Milling Cost", ev("milling_cost"))
        set_by_header("Sub Total", ev("subtotal"))
        set_by_header("P&F", ev("pf_cost"))
        set_by_header("OH & Profit", ev("oh_profit"))
        set_by_header("Rejection Cost", ev("rejection_cost"))
        set_by_header("Price/Each In INR", ev("price_each_inr"))
        set_by_header("Price/Each In Currency", ev("price_each_currency"))
        # MOQ Cost formula handled by _apply_row_formulas (Price/Each × Qty/MOQ)
        set_by_header("RM Contribution %", ev("rm_contribution_pct"))
        set_by_header("Exchange Rate", ev("exchange_rate_used"))
        set_by_header("Annual Potential", ev("annual_potential"))

    # Highlight the new row in green
    for col in range(1, (ws.max_column or 1) + 1):
        cell = ws.cell(row, col)
        cell.fill = NEW_ROW_HIGHLIGHT
        cell.border = BOLD_BORDER

    # IMPORTANT: Unhide the data row (template may have rows hidden by default)
    ws.row_dimensions[row].hidden = False
    print(f"[Excel NewFile] Data written to row {row}, unhidden")

    # Inject live Excel formulas so manual edits auto-cascade
    _apply_row_formulas(ws, row, header_map)

    # Apply header formatting
    for col in range(1, (ws.max_column or 1) + 1):
        header_cell = ws.cell(header_row, col)
        header_cell.font = STANDARD_FONT_BOLD
        header_cell.alignment = STANDARD_ALIGNMENT
        header_cell.border = BOLD_BORDER

    # Apply data validations
    _apply_data_validations(ws, header_row)
    
    # Add GLOSSARY sheet
    _create_glossary_sheet(wb)

    # Force Excel to fully recalculate all formulas on open
    wb.calculation.calcMode = 'auto'
    wb.calculation.fullCalcOnLoad = True

    # Save new file
    wb.save(output_path)
    
    print(f"[Excel NewFile] Added Part No '{part_no}' to new file: {output_path.name}")

    return ExcelWriteResult(
        output_path=output_path,
        sheet_name=used_sheet,
        header_row=header_row,
        row_written=row,
    )


