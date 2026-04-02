"""LLM-powered PDF analysis endpoints.

POST /api/v1/llm/pdf/analyze
  - Accepts a PDF upload (no job required).
  - Runs the two-agent pipeline (ExtractorAgent → ValidatorAgent).
  - Returns structured JSON with extracted specs + validation report.

POST /api/v1/llm/jobs/{job_id}/llm-analyze
  - Uses the already-uploaded source.pdf stored under an existing job.
  - Saves result to outputs/llm_analysis.json.

GET /api/v1/llm/jobs/{job_id}/llm-analysis
  - Returns cached analysis result (from last run).

GET /api/v1/llm/jobs/{job_id}/llm-analysis/export-excel
  - Downloads an Excel (.xlsx) summary of the last analysis result.
"""

import io
import json
import tempfile
from datetime import datetime, timezone
from pathlib import Path

from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse

from app.services import pdf_llm_pipeline
from app.services.job_service import JobService
from app.storage.file_storage import FileStorage

router = APIRouter()
job_service = JobService()
file_storage = FileStorage()

_RESULT_FILENAME = "llm_analysis.json"
_CORRECTIONS_FILENAME = "corrections.json"


# ---------------------------------------------------------------------------
# Stand-alone upload + analyze
# ---------------------------------------------------------------------------

@router.post("/pdf/analyze")
async def analyze_pdf_upload(file: UploadFile = File(...)):
    """Upload a PDF and run the two-agent LLM analysis pipeline.

    Returns:
    - ``extracted`` – Agent 1 structured specs (OD, ID, length, material …)
    - ``validation`` – Agent 2 cross-check report with confidence scores
    - ``code_issues`` – programmatic rule violations
    - ``valid`` – True only when the LLM says ACCEPT **and** no code issues
    """
    if not file.filename or not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="File must be a PDF")

    content = await file.read()
    if len(content) == 0:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        tmp_path = Path(tmp.name)
        tmp_path.write_bytes(content)

    try:
        result = pdf_llm_pipeline.run_pipeline(tmp_path)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except RuntimeError as exc:
        raise HTTPException(status_code=502, detail=f"LLM pipeline error: {exc}")
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Unexpected error: {exc}")
    finally:
        tmp_path.unlink(missing_ok=True)

    return result


# ---------------------------------------------------------------------------
# Analyze already-uploaded job PDF
# ---------------------------------------------------------------------------

@router.post("/jobs/{job_id}/llm-analyze")
async def analyze_job_pdf(job_id: str):
    """Run the two-agent LLM analysis on an existing job's source.pdf.

    The PDF must have been previously uploaded via
    ``POST /api/v1/jobs/{job_id}/pdf/upload``.
    """
    try:
        job_service.get_job(job_id)
    except HTTPException:
        raise HTTPException(status_code=404, detail="Job not found")

    inputs_path = file_storage.get_inputs_path(job_id)
    pdf_path = inputs_path / "source.pdf"

    if not pdf_path.exists():
        raise HTTPException(
            status_code=404,
            detail="No PDF uploaded for this job. Call /pdf/upload first.",
        )

    try:
        result = pdf_llm_pipeline.run_pipeline(pdf_path)
    except RuntimeError as exc:
        _err = str(exc)
        _is_rl = "429" in _err or "rate limit" in _err.lower()
        _rl_info = getattr(exc, "rate_limit_info", None)
        _stub = {
            "error": _err,
            "error_type": "rate_limit" if _is_rl else "pipeline_error",
            "rate_limit_info": _rl_info,
            "extracted": {},
            "validation": {
                "recommendation": "REVIEW",
                "overall_confidence": 0.0,
                "fields": {},
                "cross_checks": [
                    f"LLM analysis unavailable — Gemini 429 rate limit. "
                    f"retry_after={(_rl_info or {}).get('retry_after_s')}s  "
                    f"remaining_requests={(_rl_info or {}).get('remaining_requests')}  "
                    f"remaining_tokens={(_rl_info or {}).get('remaining_tokens')}. "
                    "Wait and try again."
                    if _is_rl else f"Pipeline error: {_err}"
                ],
            },
            "code_issues": [],
            "valid": False,
        }
        outputs_path = file_storage.get_outputs_path(job_id)
        outputs_path.mkdir(parents=True, exist_ok=True)
        (outputs_path / _RESULT_FILENAME).write_text(
            __import__("json").dumps(_stub, indent=2), encoding="utf-8"
        )
        raise HTTPException(status_code=502, detail=f"LLM pipeline error: {exc}")
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Unexpected error: {exc}")

    # Cache result for later retrieval / Excel export
    outputs_path = file_storage.get_outputs_path(job_id)
    outputs_path.mkdir(parents=True, exist_ok=True)
    (outputs_path / _RESULT_FILENAME).write_text(json.dumps(result, indent=2), encoding="utf-8")

    return result


# ---------------------------------------------------------------------------
# Get cached analysis result
# ---------------------------------------------------------------------------

@router.get("/jobs/{job_id}/llm-analysis")
async def get_job_llm_analysis(job_id: str):
    """Return the cached LLM analysis result from the last run.

    Returns ``{"available": false}`` (HTTP 200) when no analysis has been run yet,
    so the frontend can poll/auto-load without triggering a browser console error.
    """
    import traceback as _tb
    try:
        # Check result file first — serve it regardless of job registry state
        outputs_path = file_storage.get_outputs_path(job_id)
        result_path = outputs_path / _RESULT_FILENAME

        if result_path.exists():
            try:
                # Use utf-8-sig to handle files that may have been written with a BOM
                data = json.loads(result_path.read_text(encoding="utf-8-sig"))
                data["available"] = True
                return data
            except Exception as read_exc:
                raise HTTPException(
                    status_code=500,
                    detail=f"Failed to read analysis file: {read_exc}",
                )

        # File doesn't exist — validate job exists before returning available:false
        try:
            job_service.get_job(job_id)
        except HTTPException:
            raise HTTPException(status_code=404, detail="Job not found")

        return {"available": False}

    except HTTPException:
        raise
    except Exception as exc:
        _tb.print_exc()
        raise HTTPException(status_code=500, detail=f"Internal error: {exc}")


# ---------------------------------------------------------------------------
# Export cached result to Excel
# ---------------------------------------------------------------------------

@router.get("/jobs/{job_id}/llm-analysis/export-excel")
async def export_job_llm_analysis_excel(job_id: str):
    """Download an Excel summary of the last LLM analysis result."""
    try:
        job_service.get_job(job_id)
    except HTTPException:
        raise HTTPException(status_code=404, detail="Job not found")

    outputs_path = file_storage.get_outputs_path(job_id)
    result_path = outputs_path / _RESULT_FILENAME

    if not result_path.exists():
        raise HTTPException(
            status_code=404,
            detail="No LLM analysis found. Run POST /llm/jobs/{job_id}/llm-analyze first.",
        )

    result = json.loads(result_path.read_text(encoding="utf-8-sig"))
    extracted = result.get("extracted", {})
    validation = result.get("validation", {})
    code_issues = result.get("code_issues", [])
    valid = result.get("valid", False)

    try:
        import openpyxl  # type: ignore
        from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed on server")

    wb = openpyxl.Workbook()

    # ---- Sheet 1: Extracted Specs ----
    ws1 = wb.active
    ws1.title = "Extracted Specs"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    ok_fill = PatternFill("solid", fgColor="D1FAE5")
    warn_fill = PatternFill("solid", fgColor="FEF3C7")
    bad_fill = PatternFill("solid", fgColor="FEE2E2")

    def _hdr(cell, text):
        cell.value = text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    spec_fields = [
        ("Part Number",    "part_number",   None),
        ("Part Name",      "part_name",     None),
        ("Material",       "material",      None),
        ("Quantity",       "quantity",      None),
        # OD
        ("--- OD Dimensions ---", None,     None),
        ("Finish OD (in)", "od_in",         "in"),
        ("MAX OD (in)",    "max_od_in",     "in"),
        # ID
        ("--- ID Dimensions ---", None,     None),
        ("Finish ID (in)", "id_in",         "in"),
        ("MAX ID (in)",    "max_id_in",     "in"),
        # Length
        ("--- Length Dimensions ---", None, None),
        ("Finish Length (in)", "length_in", "in"),
        ("MAX Length (in)","max_length_in", "in"),
        # Meta
        ("Tolerance OD",  "tolerance_od",  None),
        ("Tolerance ID",  "tolerance_id",  None),
        ("Tolerance Len", "tolerance_length", None),
        ("Finish",        "finish",         None),
        ("Revision",      "revision",       None),
    ]

    group_fill = PatternFill("solid", fgColor="DBEAFE")  # light blue for group headers

    _hdr(ws1["A1"], "Field")
    _hdr(ws1["B1"], "Extracted Value")
    _hdr(ws1["C1"], "Confidence")
    _hdr(ws1["D1"], "LLM Issue")

    field_validations = validation.get("fields", {})

    for row, (label, key, _unit) in enumerate(spec_fields, start=2):
        if key is None:  # group header row
            cell = ws1.cell(row=row, column=1, value=label)
            cell.font = Font(bold=True, color="1E3A5F")
            cell.fill = group_fill
            for col in range(2, 5):
                ws1.cell(row=row, column=col).fill = group_fill
            continue

        val = extracted.get(key)
        vinfo = field_validations.get(key, {})
        conf = vinfo.get("confidence")
        issue = vinfo.get("issue")

        ws1.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws1.cell(row=row, column=2, value=str(val) if val is not None else "\u2014")
        ws1.cell(row=row, column=3, value=round(conf, 2) if conf is not None else "\u2014")
        ws1.cell(row=row, column=4, value=issue or "")

        row_fill = ok_fill if not issue else warn_fill
        for col in range(1, 5):
            ws1.cell(row=row, column=col).fill = row_fill

    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 22
    ws1.column_dimensions["C"].width = 14
    ws1.column_dimensions["D"].width = 40

    # ---- Sheet 2: Validation Report ----
    ws2 = wb.create_sheet("Validation Report")
    _hdr(ws2["A1"], "Section")
    _hdr(ws2["B1"], "Detail")

    overall_conf = validation.get("overall_confidence")
    recommendation = validation.get("recommendation", "—")
    cross_checks = validation.get("cross_checks", [])

    summary_rows = [
        ("Overall Confidence", f"{round(overall_conf * 100)}%" if overall_conf is not None else "—"),
        ("LLM Recommendation", recommendation),
        ("Code Issues Count", str(len(code_issues))),
        ("Final Valid", "✅ YES" if valid else "❌ NO"),
    ]
    for row, (k, v) in enumerate(summary_rows, start=2):
        ws2.cell(row=row, column=1, value=k).font = Font(bold=True)
        cell = ws2.cell(row=row, column=2, value=v)
        if k == "Final Valid":
            cell.fill = ok_fill if valid else bad_fill

    offset = len(summary_rows) + 3
    ws2.cell(row=offset, column=1, value="Cross-Checks from Agent 2").font = Font(bold=True)
    for i, check in enumerate(cross_checks or ["(none)"], start=offset + 1):
        ws2.cell(row=i, column=1, value=check)

    code_offset = offset + len(cross_checks or [""]) + 2
    ws2.cell(row=code_offset, column=1, value="Code-Level Rule Violations").font = Font(bold=True)
    for i, issue in enumerate(code_issues or ["(none)"], start=code_offset + 1):
        ws2.cell(row=i, column=1, value=issue).fill = bad_fill if code_issues else ok_fill

    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_id = job_id[:8]
    filename = f"llm_analysis_{safe_id}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


# ---------------------------------------------------------------------------
# Human-in-the-Loop dimension corrections
# ---------------------------------------------------------------------------


@router.post("/jobs/{job_id}/corrections")
async def save_correction(job_id: str, payload: dict):
    """Persist a single user-corrected dimension for a job.

    Payload shape::

        { "field": "od_in", "value": 1.880, "original_value": 1.50 }

    Corrections are stored as ``outputs/corrections.json`` alongside
    ``llm_analysis.json``.  The frontend reads them back on next load so
    overrides survive page refresh.
    """
    # Use filesystem as source-of-truth so this works even when the SQLite
    # registry has been wiped / job was created in a previous server session.
    _job_dir = file_storage.get_job_path(job_id)
    if not _job_dir.exists():
        raise HTTPException(status_code=404, detail="Job not found")

    field = payload.get("field")
    if not field:
        raise HTTPException(status_code=422, detail="'field' is required")

    outputs_path = file_storage.get_outputs_path(job_id)
    outputs_path.mkdir(parents=True, exist_ok=True)
    corrections_path = outputs_path / _CORRECTIONS_FILENAME

    existing: dict = {}
    if corrections_path.exists():
        try:
            existing = json.loads(corrections_path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            existing = {}

    existing[field] = {
        "field": field,
        "value": payload.get("value"),
        "original_value": payload.get("original_value"),
        "corrected_at": datetime.now(timezone.utc).isoformat(),
    }
    corrections_path.write_text(json.dumps(existing, indent=2), encoding="utf-8")
    return {"saved": True, "field": field, "value": payload.get("value")}


@router.get("/jobs/{job_id}/corrections")
async def get_corrections(job_id: str):
    """Return the persisted dimension corrections map for a job.

    Returns an empty object ``{}`` when no corrections have been saved yet.
    """
    # Use filesystem as source-of-truth so this works even when the SQLite
    # registry has been wiped / job was created in a previous server session.
    _job_dir = file_storage.get_job_path(job_id)
    if not _job_dir.exists():
        raise HTTPException(status_code=404, detail="Job not found")

    outputs_path = file_storage.get_outputs_path(job_id)
    corrections_path = outputs_path / _CORRECTIONS_FILENAME

    if not corrections_path.exists():
        return {}
    try:
        return json.loads(corrections_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {}
