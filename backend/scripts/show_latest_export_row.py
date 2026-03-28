from pathlib import Path
from openpyxl import load_workbook
import json

JOB_ID = "3031776f-5510-4f0b-b935-1ea58f052b37"
exports_dir = Path("backend/data/rfq_estimation/exports") / JOB_ID
files = sorted(list(exports_dir.glob("*.xlsx")), key=lambda p: p.stat().st_mtime, reverse=True)
if not files:
    print("No export files found for job", JOB_ID)
    raise SystemExit(1)

latest = files[0]
print("Latest export:", latest)
wb = load_workbook(latest, data_only=True)
ws = wb['RFQ Details'] if 'RFQ Details' in wb.sheetnames else wb.active
keys = [ws.cell(3, c).value for c in range(1, ws.max_column+1)]

# find highlighted row
found = None
for r in range(4, ws.max_row+1):
    for c in range(1, ws.max_column+1):
        cell = ws.cell(r,c)
        sc = getattr(cell.fill, 'start_color', None)
        color = None
        if sc is not None:
            color = getattr(sc, 'rgb', None) or getattr(sc, 'indexed', None) or getattr(sc, 'auto', None)
        if color and 'E6CCFF' in str(color).upper():
            found = r
            break
    if found:
        break

if not found:
    # fallback: find row where drawing_number equals job's drawing (if available)
    for r in range(4, ws.max_row+1):
        val = ws.cell(r,3).value
        if val and str(val).strip() == "050ce0004":
            found = r
            break

if not found:
    print('No highlighted or matching row found in latest export')
    raise SystemExit(1)

rowmap = {}
for c,k in enumerate(keys, start=1):
    if k:
        rowmap[k] = ws.cell(found, c).value
print(json.dumps({
    'file': str(latest),
    'row': found,
    'values': rowmap
}, indent=2, ensure_ascii=False))
