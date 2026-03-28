from openpyxl import load_workbook
p='backend/data/rfq_estimation/exports/3031776f-5510-4f0b-b935-1ea58f052b37/autofill_3031776f-5510-4f0b-b935-1ea58f052b37_09-Mar2026_160641.xlsx'
wb=load_workbook(p, data_only=False)
ws=wb['RFQ Details'] if 'RFQ Details' in wb.sheetnames else wb.active
for r in range(1, ws.max_row+1):
    for c in range(1, ws.max_column+1):
        cell=ws.cell(r,c)
        fill = cell.fill
        try:
            color = fill.start_color.rgb
        except Exception:
            color = None
        if color and 'E6CCFF' in color:
            print('highlight_row', r)
            keys=[ws.cell(3,c).value for c in range(1, ws.max_column+1)]
            rowmap={}
            for c,k in enumerate(keys, start=1):
                if k:
                    rowmap[k]=ws.cell(r,c).value
            import json
            print(json.dumps(rowmap, indent=2, ensure_ascii=False))
            raise SystemExit
print('no highlight found')
