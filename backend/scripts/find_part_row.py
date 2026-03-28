from openpyxl import load_workbook
p='backend/data/rfq_estimation/exports/3031776f-5510-4f0b-b935-1ea58f052b37/autofill_3031776f-5510-4f0b-b935-1ea58f052b37_09-Mar2026_160641.xlsx'
wb=load_workbook(p, data_only=True)
ws=wb['RFQ Details'] if 'RFQ Details' in wb.sheetnames else wb.active
target='3031776f-5510-4f0b-b935-1ea58f052b37'
rows=[]
for r in range(1, ws.max_row+1):
    v=ws.cell(r,2).value
    if v and str(v).strip()==target:
        # collect whole row mapped by key
        keys=[ws.cell(3,c).value for c in range(1, ws.max_column+1)]
        rowmap={}
        for c,k in enumerate(keys, start=1):
            if k:
                rowmap[k]=ws.cell(r,c).value
        print(rowmap)
        break
else:
    print('not found')
