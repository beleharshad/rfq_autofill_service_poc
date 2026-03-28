from openpyxl import load_workbook
import json
p='backend/data/rfq_estimation/exports/3031776f-5510-4f0b-b935-1ea58f052b37/autofill_3031776f-5510-4f0b-b935-1ea58f052b37_09-Mar2026_160641.xlsx'
wb=load_workbook(p, data_only=True)
ws=wb['RFQ Details'] if 'RFQ Details' in wb.sheetnames else wb.active
cells=[]
for r in range(1, ws.max_row+1):
    for c in range(1, ws.max_column+1):
        v=ws.cell(r,c).value
        if v is not None and (not (isinstance(v,str) and v.strip()=="")):
            cells.append({'r':r,'c':c,'v':v})
print(json.dumps(cells[:400], default=str, indent=2, ensure_ascii=False))
