import requests
import json
from pathlib import Path
from openpyxl import load_workbook

JOB_ID = "3031776f-5510-4f0b-b935-1ea58f052b37"
API = "http://127.0.0.1:8000"

def fetch_autofill():
    body = {
        "rfq_id": "TEST123",
        "part_no": JOB_ID,
        "source": {"job_id": JOB_ID},
        "tolerances": {"rm_od_allowance_in": 0.1, "rm_len_allowance_in": 0.25},
        "cost_inputs": {
            "rm_rate_per_kg": 100.0,
            "turning_rate_per_min": 4.0,
            "vmc_rate_per_min": 7.5,
            "roughing_cost": 0.0,
            "inspection_cost": 10.0,
            "others_cost": 0.0,
            "material_density_kg_m3": 7850.0,
            "pf_pct": 0.03,
            "oh_profit_pct": 0.15,
            "rejection_pct": 0.02,
            "currency": "USD",
            "use_live_rate": False,
            "qty_moq": 1,
            "annual_potential_qty": 0,
            "drawing_number": "DUMMY",
            "part_name": "TEST",
        },
    }
    resp = requests.post(f"{API}/api/v1/rfq/autofill?auto_export=false", json=body, timeout=120)
    resp.raise_for_status()
    return resp.json()


def find_latest_export():
    exports_dir = Path("backend/data/rfq_estimation/exports") / JOB_ID
    if not exports_dir.exists():
        raise FileNotFoundError(exports_dir)
    files = list(exports_dir.glob("*.xlsx"))
    files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return files[0]


def load_excel_row(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["RFQ Details"] if "RFQ Details" in wb.sheetnames else wb.active

    # try to find highlighted row (E6CCFF) or match drawing_number == JOB_ID
    keys = [ws.cell(3, c).value for c in range(1, ws.max_column + 1)]

    # attempt highlight detection (fill color can be stored differently; check a few rows)
    for r in range(4, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            sc = getattr(cell.fill, 'start_color', None)
            color = None
            if sc is not None:
                color = getattr(sc, 'rgb', None) or getattr(sc, 'indexed', None) or getattr(sc, 'auto', None)
            if color and 'E6CCFF' in str(color).upper():
                # build row map
                rowmap = {}
                for ci, k in enumerate(keys, start=1):
                    if k:
                        rowmap[k] = ws.cell(r, ci).value
                return rowmap

    # fallback: find row where drawing_number or part_no matches JOB_ID
    for r in range(4, ws.max_row + 1):
        val = ws.cell(r, 3).value
        if val and str(val).strip() == JOB_ID:
            rowmap = {}
            for ci, k in enumerate(keys, start=1):
                if k:
                    rowmap[k] = ws.cell(r, ci).value
            return rowmap

    raise RuntimeError("Could not locate exported row in workbook")


def compare(autofill, excel_row):
    fields = autofill.get('fields', {})
    estimate = autofill.get('estimate', {})

    mapping = {
        'finish_od': ('fields', 'finish_od_in'),
        'finish_id': ('fields', 'finish_id_in'),
        'finish_length': ('fields', 'finish_len_in'),
        'finish_od_mm': ('fields', 'finish_od_mm'),
        'finish_id_mm': ('fields', 'finish_id_mm'),
        'finish_length_mm': ('fields', 'finish_len_mm'),
        'od': ('fields', 'rm_od_in'),
        'rm_length': ('fields', 'rm_len_in'),
        'price_each_inr': ('estimate', 'price_each_inr'),
        'price_in_currency': ('estimate', 'price_each_currency'),
    }

    diffs = []
    for key, (block, name) in mapping.items():
        if block == 'fields':
            val = fields.get(name, {}).get('value')
        else:
            val = estimate.get(name, {}).get('value')

        excel_val = excel_row.get(key)

        # normalize numbers
        try:
            if val is None:
                equal = excel_val is None or excel_val == " " or excel_val == ""
            else:
                equal = abs(float(val) - float(excel_val)) < 0.01
        except Exception:
            equal = str(val) == str(excel_val)

        diffs.append({
            'key': key,
            'autofill': val,
            'excel': excel_val,
            'match': equal,
        })

    return diffs


def main():
    print("Fetching autofill response from API...")
    autofill = fetch_autofill()
    print("Autofill fetched.")

    xlsx = find_latest_export()
    print("Using Excel:", xlsx)
    excel_row = load_excel_row(xlsx)
    diffs = compare(autofill, excel_row)
    print(json.dumps(diffs, indent=2, ensure_ascii=False))


if __name__ == '__main__':
    main()
