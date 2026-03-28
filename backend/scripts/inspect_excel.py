from openpyxl import load_workbook
import json
p='backend/data/rfq_estimation/exports/3031776f-5510-4f0b-b935-1ea58f052b37/autofill_3031776f-5510-4f0b-b935-1ea58f052b37_09-Mar2026_160641.xlsx'
wb=load_workbook(p, data_only=True)
ws=wb['RFQ Details'] if 'RFQ Details' in wb.sheetnames else wb.active
headers=[ws.cell(2,c).value for c in range(1, ws.max_column+1)]
keys=[ws.cell(3,c).value for c in range(1, ws.max_column+1)]
cols=[]
for c,h,k in zip(range(1, ws.max_column+1), headers, keys):
    cols.append({'col':c,'header':h,'key':k})
print(json.dumps(cols, indent=2, ensure_ascii=False))
