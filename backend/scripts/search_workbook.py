from openpyxl import load_workbook
import sys
import json
p='backend/data/rfq_estimation/exports/3031776f-5510-4f0b-b935-1ea58f052b37/autofill_3031776f-5510-4f0b-b935-1ea58f052b37_09-Mar2026_160641.xlsx'
wb=load_workbook(p, data_only=True)
found=False
for ws in wb.worksheets:
    for r in range(1, ws.max_row+1):
        for c in range(1, ws.max_column+1):
            v=ws.cell(r,c).value
            if v and str(v).strip()== '3031776f-5510-4f0b-b935-1ea58f052b37':
                print({'sheet':ws.title,'r':r,'c':c,'v':v})
                found=True
if not found:
    print('not found')
