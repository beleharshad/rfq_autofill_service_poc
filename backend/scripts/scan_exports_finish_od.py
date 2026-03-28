from pathlib import Path
from openpyxl import load_workbook
p=Path('backend/data/rfq_estimation/exports/3031776f-5510-4f0b-b935-1ea58f052b37')
if not p.exists():
    print('No exports dir at', p)
    raise SystemExit(0)
for f in sorted(p.glob('*.xlsx')):
    try:
        wb=load_workbook(f,data_only=True)
        ws=wb['RFQ Details'] if 'RFQ Details' in wb.sheetnames else wb.active
        keys=[ws.cell(3,c).value for c in range(1,ws.max_column+1)]
        # find finish_od column
        col=None
        for i,k in enumerate(keys, start=1):
            if k and str(k).strip()=='finish_od':
                col=i
                break
        values=[]
        if col:
            for r in range(4, ws.max_row+1):
                v=ws.cell(r,col).value
                if v is not None:
                    values.append((r,v))
        print(f.name, '->', values[:10])
    except Exception as e:
        print('ERR', f.name, e)
