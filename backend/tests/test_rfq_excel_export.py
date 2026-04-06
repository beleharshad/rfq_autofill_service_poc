from pathlib import Path

import openpyxl

from app.services.rfq_excel_export_service import HEADER_ROW, write_autofill_to_new_file


def test_new_file_export_inserts_drilling_and_milling_columns(tmp_path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RFQ Details"

    headers = [
        "Sr.No",
        "Part No",
        "Finish OD (Inch)",
        "Finish ID (Inch)",
        "Finish Length (Inch)",
        "Finish OD (MM)",
        "Finish ID (MM)",
        "Finish Length (MM)",
        "RM OD (Inch)",
        "RM ID (Inch)",
        "Length (Inch)",
        "RM Weight Kg",
        "RM Rate",
        "Material Cost",
        "Roughing Cost",
        "Turning Time In Min",
        "Turning cost",
        "VMC Time In Min",
        "VMC cost",
        "Currency",
        "Sub Total",
        "P&F",
        "OH & Profit",
        "Rejection Cost",
        "Price/Each In INR",
        "Price/Each In Currency",
        "RM Contribution %",
        "Exchange Rate",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(HEADER_ROW, col).value = header

    wb.save(template_path)

    result = write_autofill_to_new_file(
        template_path=template_path,
        output_path=output_path,
        part_no="MIXER-001",
        autofill_response={
            "fields": {
                "finish_od_in": {"value": 6.811},
                "finish_id_in": {"value": 0.118},
                "finish_len_in": {"value": 6.1},
                "finish_od_mm": {"value": 172.999},
                "finish_id_mm": {"value": 2.997},
                "finish_len_mm": {"value": 154.94},
                "rm_od_in": {"value": 6.911},
                "rm_id_in": {"value": 0.0},
                "rm_len_in": {"value": 6.45},
            },
            "estimate": {
                "rm_weight_kg": {"value": 31.136},
                "material_cost": {"value": 2802.24},
                "roughing_cost": {"value": 0.0},
                "turning_minutes": {"value": 38.735},
                "turning_cost": {"value": 290.512},
                "vmc_minutes": {"value": 12.0},
                "vmc_cost": {"value": 90.0},
                "drilling_minutes": {"value": 12.0},
                "drilling_cost": {"value": 90.0},
                "milling_minutes": {"value": 0.0},
                "milling_cost": {"value": 0.0},
                "special_process_cost": {"value": 25.0},
                "others_cost": {"value": 11.0},
                "inspection_cost": {"value": 10.0},
                "subtotal": {"value": 3182.752},
                "pf_cost": {"value": 95.483},
                "oh_profit": {"value": 636.55},
                "rejection_cost": {"value": 63.655},
                "price_each_inr": {"value": 3978.44},
                "price_each_currency": {"value": 48.517},
                "rm_contribution_pct": {"value": 70.432},
                "exchange_rate_used": {"value": 82.0},
                "annual_potential": {"value": 4851.7},
            },
        },
        cost_inputs={"rm_rate_per_kg": 90.0, "qty_moq": 5, "annual_potential_qty": 100, "currency": "USD"},
        sheet_name="RFQ Details",
    )

    assert Path(result.output_path).exists()

    out_wb = openpyxl.load_workbook(output_path, data_only=False)
    out_ws = out_wb["RFQ Details"]
    header_map = {
        str(out_ws.cell(HEADER_ROW, col).value).strip(): col
        for col in range(1, out_ws.max_column + 1)
        if out_ws.cell(HEADER_ROW, col).value
    }

    assert "Drilling Time In Min" in header_map
    assert "Drilling Cost" in header_map
    assert "Milling Time In Min" in header_map
    assert "Milling Cost" in header_map
    assert "Special Process Cost" in header_map
    assert "Others" in header_map
    assert "Inspection and testing Cost" in header_map
    assert "Exchange Rate" in header_map
    assert "Annual Potential" in header_map
    assert "MOQ Cost" in header_map
    assert header_map["Turning cost"] < header_map["Drilling Time In Min"] < header_map["Drilling Cost"] < header_map["Milling Time In Min"] < header_map["Milling Cost"]
    assert out_ws.cell(3, header_map["Drilling Time In Min"]).value == 12.0
    assert out_ws.cell(3, header_map["Drilling Cost"]).value == 90.0
    assert out_ws.cell(3, header_map["Special Process Cost"]).value == 25.0
    assert out_ws.cell(3, header_map["Others"]).value == 11.0
    assert out_ws.cell(3, header_map["Inspection and testing Cost"]).value == 10.0
    assert out_ws.cell(3, header_map["Exchange Rate"]).value == 82.0
    assert out_ws.cell(3, header_map["Annual Potential"]).value == 4851.7